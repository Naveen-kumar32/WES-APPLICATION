from django.shortcuts import render

# Create your views here.
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse
from io import BytesIO 
from xhtml2pdf import pisa  
from django.http import HttpResponse
import csv
from django.http import JsonResponse
import os
from django.db.models import Q
from django.conf import settings
from decimal import Decimal
from django.db import transaction
from django_pivot.pivot import pivot
from collections import defaultdict
from PIL import Image
import base64
from django.contrib import messages
from .models import DHL,CommercialInvoice,DoDnNumber,CommercialPacl,DoNumber,Ordertracking
from .forms import DHLForm,commercialinvoiceForm,dodnnumberForm,commercialpaclForm,donumberForm,OrdertrackingForm
from django.db.models import F,ExpressionWrapper, IntegerField
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import FileResponse
import re
from openpyxl.styles import Font, PatternFill

def logindex(request):

    return render(request, 'logindex.html')

def dhl(request):
    filter_FromCountry = request.GET.get('filter_FromCountry')
    filter_ToCountry = request.GET.get('filter_ToCountry')
    filter_Weight_Kg = request.GET.get('filter_Weight_Kg')

    test1 = DHL.objects.all().order_by('-id')

    if filter_FromCountry:
        test1 = test1.filter(From_Country=filter_FromCountry)
    if filter_ToCountry:
        test1 = test1.filter(To_Country=filter_ToCountry)
    if filter_Weight_Kg:
        test1 = test1.filter(Weight_Kg=filter_Weight_Kg)

    filtered_data = list(test1.values())
    total_rows = len(filtered_data)
    context = {
        'test1': test1,
        'filter_FromCountry': filter_FromCountry,
        'filter_ToCountry': filter_ToCountry,
        'filter_Weight_Kg': filter_Weight_Kg,
        'total_rows':int(total_rows),
    }
    # return JsonResponse({'total_rows': total_rows})
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        return JsonResponse(filtered_data, safe=False)


    return render(request, 'DHL/dhl.html', context)


def get_dropdown_options1(request):

    selected_fromcountry = request.GET.get('From_Country')
    test1 = DHL.objects.all()
    if selected_fromcountry:
        invoices = test1.filter(From_Country=selected_fromcountry)
        to_country_options = list(invoices.values_list('To_Country', flat=True).distinct())
        weight1_options = list(invoices.values_list('Weight_Kg', flat=True).distinct())
    else:
        to_country_options = []
        weight1_options = []

    response_data = {
        'to_country_options': to_country_options,
        'weight1_options': weight1_options,
    }

    return JsonResponse(response_data)


def get_dropdown_optionstocountry(request):

    selected_fromcountry = request.GET.get('From_Country')
    selected_tocountry = request.GET.get('To_Country')
    test1 = DHL.objects.all()
    if selected_tocountry:
        invoices = test1.filter(From_Country=selected_fromcountry,To_Country=selected_tocountry)
        weight1_options = list(invoices.values_list('Weight_Kg', flat=True).distinct())
    else:

        weight1_options = []

    response_data = {
        'weight1_options': weight1_options,
    }

    return JsonResponse(response_data)


def dhlindex(request):
    if request.method == "POST":
        form = DHLForm(request.POST)
        if form.is_valid():
            form.save()
            filter_FromCountry  = form.cleaned_data['From_Country']
            filter_ToCountry  = form.cleaned_data['To_Country']
            filter_Weight_Kg = form.cleaned_data['Weight_Kg']
            url = reverse('testcase:filter')
            url += f'?filter_FromCountry={filter_FromCountry}&filter_ToCountry={filter_ToCountry}&filter_Weight_Kg={filter_Weight_Kg}'
            return HttpResponseRedirect(url)
    else:
        form = DHLForm()


    if request.GET.get('filter_FromCountry'):
        filter_FromCountry1 = request.GET.get('filter_FromCountry')
        # Perform any filtering or processing based on the filter_ClientName value

        if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
            selectedFromCountry = request.GET.get('selectedFromCountry')
            print(f"Received selected_client_name: { selectedFromCountry }")
            # Return a JSON response if it's an AJAX request
            return JsonResponse({'message': 'Filtered data'})
    test1 = DHL.objects.all()


    filter_FromCountry = request.GET.get('filter_FromCountry')
    filter_ToCountry = request.GET.get('filter_ToCountry')
    filter_Weight_Kg = request.GET.get('filter_Weight_Kg')

    print(filter_FromCountry)

    if filter_FromCountry:
        test1 = test1.filter(From_Country=filter_FromCountry)
    if filter_ToCountry:
        test1 = test1.filter(To_Country=filter_ToCountry)
    if filter_Weight_Kg:
        test1 = test1.filter(Weight_Kg=filter_Weight_Kg)

    unique_FromCountry = DHL.objects.order_by('From_Country').values_list('From_Country', flat=True).distinct()
    unique_ToCountry = DHL.objects.order_by('To_Country').values_list('To_Country', flat=True).distinct()
    unique_Weight_Kg = DHL.objects.order_by('Weight_Kg').values_list('Weight_Kg', flat=True).distinct()

    print('hi')
    context = {
        'form': form,
        'test1': test1,
        'filter_FromCountry': filter_FromCountry,
        'filter_ToCountry': filter_ToCountry,
        'filter_Weight_Kg': filter_Weight_Kg,
        'unique_FromCountry': unique_FromCountry,
        'unique_ToCountry': unique_ToCountry,
        'unique_Weight_Kg': unique_Weight_Kg,
    }
    return render(request, 'DHL/dhlindex.html', context)


def dhlexcel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Full Data"
    testdata_objects = DHL.objects.all()

    # Define the desired fields for the CSV file
    fields = ['id'	,'WES_Ref',	'PO_NO'	,'Client_Name'	,'Ship_name',	'Client_Invoice_No'	,'INVOICE_DATE',	'Client_Frieght_Cost'	,'Client_Freight_Currency'
              	,'Client_Freight_Cost_in_SGD'	,'DHL_Invoice_number'	,'AWB_NUMBER'	,'AMOUNT_INR'	,'DHL_AMOUNT_SGD',	'DHL_DUTY_TAX'	,'Invoice_date2',
                'Due_Date'	,'Status',	'Paid_date',	'Transaction_number'	,'From_Country'	, 'To_Country',	'Weight_Kg','Dimension_Volume', 'Dimension_CM' ,
                'Profit_and_Loss'	,'Remarks']

    for col_num, field in enumerate(fields, 1):
        # col_letter = get_column_letter(col_num)
        cell = ws.cell(row=1, column=col_num, value=field)
        cell.font = openpyxl.styles.Font(bold=True)

    # Write the data rows
    for row_num, data in enumerate(testdata_objects, 2):
        for col_num, field in enumerate(fields, 1):
            cell = ws.cell(row=row_num, column=col_num, value=getattr(data, field))

    file_path = "full_data.xlsx"
    wb.save(file_path)

    response = FileResponse(open(file_path, 'rb'), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="full_data.xlsx'

    return response

def dhlfullexcel(request):
    # Get all Testdata objects without applying any filters
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Full Data"
    testdata_objects = DHL.objects.all()

    # Define the fields for the CSV file
    fields = ['id'	,'WES_Ref',	'PO_NO'	,'Client_Name'	,'Ship_name',	'Client_Invoice_No'	,'INVOICE_DATE',	'Client_Frieght_Cost'	,'Client_Freight_Currency'
              	,'Client_Freight_Cost_in_SGD'	,'DHL_Invoice_number'	,'AWB_NUMBER'	,'AMOUNT_INR'	,'DHL_AMOUNT_SGD',	'DHL_DUTY_TAX'	,'Invoice_date2',
                'Due_Date'	,'Status',	'Paid_date',	'Transaction_number'	,'From_Country'	, 'To_Country',	'Weight_Kg','Dimension_Volume', 'Dimension_CM' ,
                'Profit_and_Loss'	,'Remarks']
    # Write the headers
    for col_num, field in enumerate(fields, 1):
        # col_letter = get_column_letter(col_num)
        cell = ws.cell(row=1, column=col_num, value=field)
        cell.font = openpyxl.styles.Font(bold=True)

    # Write the data rows
    for row_num, data in enumerate(testdata_objects, 2):
        for col_num, field in enumerate(fields, 1):
            cell = ws.cell(row=row_num, column=col_num, value=getattr(data, field))

    file_path = "full_data.xlsx"
    wb.save(file_path)

    response = FileResponse(open(file_path, 'rb'), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="full_data.xlsx'

    return response


def adddhl(request):
    if request.method == "POST":
        form = DHLForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                return redirect('Logistics:dhlindex')
            except Exception as e:
                return HttpResponse(f"An error occurred: {str(e)}")
        else:
            return HttpResponse("Form is not valid.")
    else:
        form = DHLForm()

    Client_Name = DHL.objects.exclude(Client_Name__exact='').order_by('Client_Name').values_list('Client_Name', flat=True).distinct()
    Ship_name = DHL.objects.exclude(Ship_name__exact='').order_by('Ship_name').values_list('Ship_name', flat=True).distinct()
    Status  = DHL.objects.exclude(Status__exact='').order_by('Status').values_list('Status', flat=True).distinct()
    From_Country  = DHL.objects.exclude(From_Country__exact='').order_by('From_Country').values_list('From_Country', flat=True).distinct()
    To_Country  = DHL.objects.exclude(To_Country__exact='').order_by('To_Country').values_list('To_Country', flat=True).distinct()
    Client_Freight_Currency = DHL.objects.exclude(Client_Freight_Currency__exact='').order_by('Client_Freight_Currency').values_list('Client_Freight_Currency', flat=True).distinct()
    last_id = DHL.objects.last().id if DHL.objects.exists() else 0
    context = {'last_id': last_id + 1} 
      

    context = {
        'ClientName_options': Client_Name,
        'shipname_options': Ship_name,
        'newstatus_options': Status,
        'From_Country_options': From_Country,
        'To_Country_options': To_Country,
        'form': form,
        'Client_Freight_Currency_options' : Client_Freight_Currency,
        'last_id': last_id + 1

    }
    return render(request, 'DHL/adddhl.html', context)
    
def addnewdhl(request):
    if request.method == "POST":
        form = DHLForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                return redirect('Logistics:dhlindex')
            except Exception as e:
                return HttpResponse(f"An error occurred: {str(e)}")
        else:
            return HttpResponse("Form is not valid.")
    else:
        form = DHLForm()

    Client_Name = DHL.objects.exclude(Client_Name__exact='').order_by('Client_Name').values_list('Client_Name', flat=True).distinct()
    Ship_name = DHL.objects.exclude(Ship_name__exact='').order_by('Ship_name').values_list('Ship_name', flat=True).distinct()
    Status  = DHL.objects.exclude(Status__exact='').order_by('Status').values_list('Status', flat=True).distinct()
    From_Country  = DHL.objects.exclude(From_Country__exact='').order_by('From_Country').values_list('From_Country', flat=True).distinct()
    To_Country  = DHL.objects.exclude(To_Country__exact='').order_by('To_Country').values_list('To_Country', flat=True).distinct()
    Client_Freight_Currency = DHL.objects.exclude(Client_Freight_Currency__exact='').order_by('Client_Freight_Currency').values_list('Client_Freight_Currency', flat=True).distinct()
    last_id = DHL.objects.last().id if DHL.objects.exists() else 0
    context = {'last_id': last_id + 1} 
      

    context = {
        'ClientName_options': Client_Name,
        'shipname_options': Ship_name,
        'newstatus_options': Status,
        'From_Country_options': From_Country,
        'To_Country_options': To_Country,
        'form': form,
        'Client_Freight_Currency_options' : Client_Freight_Currency,
        'last_id': last_id + 1

    }
    return render(request, 'DHL/addnewdhl.html',context)


def updatedhl(request, data_id):
    data = get_object_or_404(DHL, pk=data_id)
    if request.method == 'POST': 
        form = DHLForm(request.POST, instance=data)

        if form.is_valid():
            form.save()
            return redirect('Logistics:dhlindex')
    else:
        form = DHLForm(instance=data)
        
    shipname_data = DHL.objects.exclude(Ship_name__exact='').order_by('Ship_name').values_list('Ship_name', flat=True).distinct()
    Client_Name = DHL.objects.exclude(Client_Name__exact='').order_by('Client_Name').values_list('Client_Name', flat=True).distinct()
    To_Country = DHL.objects.exclude(To_Country__exact='').order_by('To_Country').values_list('To_Country', flat=True).distinct()
    From_Country = DHL.objects.exclude(From_Country__exact='').order_by('From_Country').values_list('From_Country', flat=True).distinct()
   
    
    context = {
        'form': form,
        'data_id': data_id,
        'shipname_options': shipname_data,
        'ClientName_options' : Client_Name,
        'From_Country':From_Country,
        'To_Country' : To_Country
        
    }
    return render(request, 'DHL/updatedhl.html', context)


def commercialinvoice(request):
 
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        return JsonResponse(filtered_data, safe=False)
    
    
    return render(request, 'commercialinvoice/commercialinvoice.html')

def addci(request):
    if request.method == "POST":
        form = commercialinvoiceForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                return redirect('Logistics:commercialinvoice')
            except Exception as e:
                return HttpResponse(f"An error occurred: {str(e)}")
        else:
            return HttpResponse("Form is not valid.")
    else:
        form = commercialinvoiceForm()

    last_id = CommercialInvoice.objects.last().ID if CommercialInvoice.objects.exists() else 0
    VESSEL_NAME = CommercialInvoice.objects.exclude(VESSEL_NAME__exact='').order_by('VESSEL_NAME').values_list('VESSEL_NAME', flat=True).distinct()

    context = {
        
        'last_id': last_id + 1,
        'form': form,
        'VESSEL_NAME_options' : VESSEL_NAME
    
    
    } 
      
    return render(request, 'commercialinvoice/addci.html',context)

def addnewci(request):
    if request.method == "POST":
        form = commercialinvoiceForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                return redirect('Logistics:commercialinvoice')
            except Exception as e:
                return HttpResponse(f"An error occurred: {str(e)}")
        else:
            return HttpResponse("Form is not valid.")
    else:
        form = commercialinvoiceForm()

    last_id = CommercialInvoice.objects.last().ID if CommercialInvoice.objects.exists() else 0

    context = {
        
        'last_id': last_id + 1,
        'form': form
    
    } 
      
    return render(request, 'commercialinvoice/addnewci.html',context)
   
def commercialinvoiceci(request):
      
    
    test1 = CommercialInvoice.objects.all().order_by('-ID')
   

    filtered_data = list(test1.values())    
    total_rows = len(filtered_data)
    context = {
        'test1': test1,
       
        'total_rows':int(total_rows),
    }
    # return JsonResponse({'total_rows': total_rows})
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        return JsonResponse(filtered_data, safe=False)
    
    return render(request,'commercialinvoice/commercialinvoiceci.html', context)            

def dodnnumber(request):
 
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        return JsonResponse(filtered_data, safe=False)
    
    
    return render(request, 'dodnnumber/dodnnumber.html')
def adddn(request):
      
    if request.method == "POST":
        form = dodnnumberForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                return redirect('Logistics:dodnnumber')
            except Exception as e:
                return HttpResponse(f"An error occurred: {str(e)}")
        else:
            return HttpResponse("Form is not valid.")
    else:
        form = dodnnumberForm()

    last_id = DoDnNumber.objects.last().ID if DoDnNumber.objects.exists() else 0
    VESSEL_NAME = DoDnNumber.objects.exclude(VESSEL_NAME__exact='').order_by('VESSEL_NAME').values_list('VESSEL_NAME', flat=True).distinct()


    context = {
        
        'last_id': last_id + 1,
        'form': form,
        'VESSEL_NAME_options' : VESSEL_NAME
    
    } 
      
    return render(request, 'dodnnumber/adddn.html',context)

def addnewdn(request):
    if request.method == "POST":
        form = dodnnumberForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                return redirect('Logistics:dodnnumber')
            except Exception as e:
                return HttpResponse(f"An error occurred: {str(e)}")
        else:
            return HttpResponse("Form is not valid.")
    else:
        form = dodnnumberForm()

    last_id = DoDnNumber.objects.last().ID if DoDnNumber.objects.exists() else 0

    context = {
        
        'last_id': last_id + 1,
        'form': form
    
    
    
    } 
      
    return render(request, 'dodnnumber/addnewdn.html',context)
def dodnnumberdn(request):
      
    test1 = DoDnNumber.objects.all().order_by('-ID')
   

    filtered_data = list(test1.values())    
    total_rows = len(filtered_data)
    context = {
        'test1': test1,
       
        'total_rows':int(total_rows),
    }
    # return JsonResponse({'total_rows': total_rows})
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        return JsonResponse(filtered_data, safe=False)
    
    return render(request,'dodnnumber/dodnnumberdn.html', context)      
 

def updateci(request, data_id):
    data = get_object_or_404(CommercialInvoice, pk=data_id)  # Use CommercialInvoice model
    if request.method == 'POST':
        form = commercialinvoiceForm(request.POST, instance=data)

        if form.is_valid():
            form.save()
            return redirect('Logistics:commercialinvoiceci')
    else:
        form = commercialinvoiceForm(instance=data) 
    
    context = {
        'form': form,
        'data_id': data_id,
    }
    return render(request, 'commercialinvoice/updateci.html', context)    

def updatedn(request, data_id):
    data = get_object_or_404(DoDnNumber, pk=data_id)
    if request.method == 'POST':
        form = dodnnumberForm(request.POST, instance=data)

        if form.is_valid():
            form.save()
            return redirect('Logistics:dodnnumberdn')
    else:
        form = dodnnumberForm(instance=data) 
    
    context = {
        'form': form,
        'data_id': data_id,
       
    }
    return render(request, 'dodnnumber/updatedn.html', context)

def commercialpacl(request):
 
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        return JsonResponse(filtered_data, safe=False)
    
    
    return render(request, 'commercialpacl/commercialpacl.html')

def addcp(request):
    if request.method == "POST":
        form = commercialpaclForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                return redirect('Logistics:commercialpacl')
            except Exception as e:
                return HttpResponse(f"An error occurred: {str(e)}")
        else:
            return HttpResponse("Form is not valid.")
    else:
        form = commercialpaclForm()

    last_id = CommercialPacl.objects.last().ID if CommercialPacl.objects.exists() else 0
    VESSEL_NAME = CommercialPacl.objects.exclude(VESSEL_NAME__exact='').order_by('VESSEL_NAME').values_list('VESSEL_NAME', flat=True).distinct()

    context = {
        
        'last_id': last_id + 1,
        'form': form,
        'VESSEL_NAME_options' : VESSEL_NAME
    
    
    } 
      
    return render(request, 'commercialpacl/addcp.html',context)

def addnewcp(request):
    if request.method == "POST":
        form = commercialpaclForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                return redirect('Logistics:commercialpacl')
            except Exception as e:
                return HttpResponse(f"An error occurred: {str(e)}")
        else:
            return HttpResponse("Form is not valid.")
    else:
        form = commercialpaclForm()

    last_id = CommercialPacl.objects.last().ID if CommercialPacl.objects.exists() else 0

    context = {
        
        'last_id': last_id + 1,
        'form': form
    
    } 
      
    return render(request, 'commercialpacl/addnewcp.html',context)
   
def commercialpaclcp(request):
      
    
    test1 = CommercialPacl.objects.all().order_by('-ID')
   

    filtered_data = list(test1.values())    
    total_rows = len(filtered_data)
    context = {
        'test1': test1,
       
        'total_rows':int(total_rows),
    }
    # return JsonResponse({'total_rows': total_rows})
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        return JsonResponse(filtered_data, safe=False)
    
    return render(request,'commercialpacl/commercialpaclcp.html', context)       



def updatecp(request, data_id):
    data = get_object_or_404(CommercialPacl, pk=data_id)  # Use CommercialInvoice model
    if request.method == 'POST':
        form = commercialpaclForm(request.POST, instance=data)

        if form.is_valid():
            form.save()
            return redirect('Logistics:commercialpaclcp')
    else:
        form = commercialpaclForm(instance=data) 
    
    context = {
        'form': form,
        'data_id': data_id,
    }
    return render(request, 'commercialpacl/updatecp.html', context) 

def donumber(request):
 
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        return JsonResponse(filtered_data, safe=False)
    
    
    return render(request, 'donumber/donumber.html')

def adddo(request):
    if request.method == "POST":
        form = donumberForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                return redirect('Logistics:donumber')
            except Exception as e:
                return HttpResponse(f"An error occurred: {str(e)}")
        else:
            return HttpResponse("Form is not valid.")
    else:
        form = donumberForm()

    last_id = DoNumber.objects.last().ID if DoNumber.objects.exists() else 0
    VESSEL_NAME = DoNumber.objects.exclude(VESSEL_NAME__exact='').order_by('VESSEL_NAME').values_list('VESSEL_NAME', flat=True).distinct()

    context = {
        
        'last_id': last_id + 1,
        'form': form,
        'VESSEL_NAME_options' : VESSEL_NAME
    
    
    } 
      
    return render(request, 'donumber/adddo.html',context)

def addnewdo(request):
    if request.method == "POST":
        form = donumberForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                return redirect('Logistics:donumber')
            except Exception as e:
                return HttpResponse(f"An error occurred: {str(e)}")
        else:
            return HttpResponse("Form is not valid.")
    else:
        form = donumberForm()

    last_id = DoNumber.objects.last().ID if DoNumber.objects.exists() else 0

    context = {
        
        'last_id': last_id + 1,
        'form': form
    
    } 
      
    return render(request, 'donumber/addnewdo.html',context)
   
def donumberdo(request):
      
    
    test1 = DoNumber.objects.all().order_by('-ID')
   

    filtered_data = list(test1.values())    
    total_rows = len(filtered_data)
    context = {
        'test1': test1,
       
        'total_rows':int(total_rows),
    }
    # return JsonResponse({'total_rows': total_rows})
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        return JsonResponse(filtered_data, safe=False)
    
    return render(request,'donumber/donumberdo.html', context)       


def updatedo(request, data_id):
    data = get_object_or_404(DoNumber, pk=data_id)  # Use CommercialInvoice model
    if request.method == 'POST':
        form = donumberForm(request.POST, instance=data)

        if form.is_valid():
            form.save()
            return redirect('Logistics:donumberdo')
    else:
        form = donumberForm(instance=data) 
    
    context = {
        'form': form,
        'data_id': data_id,
    }
    return render(request, 'donumber/updatedo.html', context)         

def download_ci_csv(request):
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Full Data"
    testdata_objects = CommercialInvoice.objects.all()
    # Define the fields for the CSV file
    fields = ['ID','COMMERCIAL_NUMBER','PACKING_LIST_NUMBER', 'DATE', 'VESSEL_NAME', 'WES_NUMBER', 'PO_NUMBER', 'INCHARGE', 'Remarks']
    
    # Write the headers
    for col_num, field in enumerate(fields, 1):
        # col_letter = get_column_letter(col_num)
        cell = ws.cell(row=1, column=col_num, value=field)
        cell.font = openpyxl.styles.Font(bold=True) 

    # Write the data rows
    for row_num, data in enumerate(testdata_objects, 2):
        for col_num, field in enumerate(fields, 1):
            cell = ws.cell(row=row_num, column=col_num, value=getattr(data, field))

    file_path = "full_data.xlsx"
    wb.save(file_path)

    response = FileResponse(open(file_path, 'rb'), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="full_data.xlsx'

    return response


def download_dodnnumber_csv(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Full Data"
    testdata_objects = DoDnNumber.objects.all()
    # Define the fields for the CSV file
    fields = ['ID','DO_DN_NUMBER', 'DATE', 'VESSEL_NAME', 'WES_NUMBER', 'PO_NUMBER', 'INCHARGE', 'REMARK']
    
    # Write the headers
    for col_num, field in enumerate(fields, 1):
        # col_letter = get_column_letter(col_num)
        cell = ws.cell(row=1, column=col_num, value=field)
        cell.font = openpyxl.styles.Font(bold=True) 

    # Write the data rows
    for row_num, data in enumerate(testdata_objects, 2):
        for col_num, field in enumerate(fields, 1):
            cell = ws.cell(row=row_num, column=col_num, value=getattr(data, field))

    file_path = "full_data.xlsx"
    wb.save(file_path)

    response = FileResponse(open(file_path, 'rb'), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="full_data.xlsx'

    return response   

def download_commercial_Pacl_csv(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Full Data"
    testdata_objects = CommercialPacl.objects.all()
    # Define the fields for the CSV file
    # Define the fields for the CSV file
    fields = ['ID','COMMERCIAL_NUMBER', 'PACL_NUMBER', 'DATE', 'VESSEL_NAME', 'WES_NUMBER', 'PO_NUMBER', 'INCHARGE']
   
    for col_num, field in enumerate(fields, 1):
        # col_letter = get_column_letter(col_num)
        cell = ws.cell(row=1, column=col_num, value=field)
        cell.font = openpyxl.styles.Font(bold=True) 

    # Write the data rows
    for row_num, data in enumerate(testdata_objects, 2):
        for col_num, field in enumerate(fields, 1):
            cell = ws.cell(row=row_num, column=col_num, value=getattr(data, field))

    file_path = "full_data.xlsx"
    wb.save(file_path)

    response = FileResponse(open(file_path, 'rb'), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="full_data.xlsx'

    return response            

def download_Do_Number_csv(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Full Data"
    testdata_objects = DoNumber.objects.all()

    # Define the fields for the CSV file
    fields = ['ID','DO_NUMBER', 'DATE', 'VESSEL_NAME', 'WES_NUMBER', 'PO_NUMBER', 'INCHARGE']
   
    for col_num, field in enumerate(fields, 1):
        # col_letter = get_column_letter(col_num)
        cell = ws.cell(row=1, column=col_num, value=field)
        cell.font = openpyxl.styles.Font(bold=True) 

    # Write the data rows
    for row_num, data in enumerate(testdata_objects, 2):
        for col_num, field in enumerate(fields, 1):
            cell = ws.cell(row=row_num, column=col_num, value=getattr(data, field))

    file_path = "full_data.xlsx"
    wb.save(file_path)

    response = FileResponse(open(file_path, 'rb'), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="full_data.xlsx'

    return response            

def ordertracking(request):
    if request.method == "POST":
        form = OrdertrackingForm(request.POST)
        if form.is_valid():
            form.save()
            filter_Incharge = form.cleaned_data['InCharger']
            filter_Status = form.cleaned_data['Status']

            url = reverse('Logistics:filter')
            url += f'?filter_Incharge={filter_Incharge}&filter_Status={filter_Status}'
            return HttpResponseRedirect(url)
    else:
        form = OrdertrackingForm()


    if request.GET.get('filter_Incharge'):
        filter_Incharge = request.GET.get('filter_Incharge')
        # Perform any filtering or processing based on the filter_ClientName value

        if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
            # Return a JSON response if it's an AJAX request
            return JsonResponse({'message': 'Filtered data'})
    test1 = Ordertracking.objects.all().order_by('-ID')

    filter_Incharge = request.GET.get('filter_Incharge')
    filter_Status = request.GET.get('filter_Status')

    if filter_Incharge:
        test1 = test1.filter(InCharger=filter_Incharge)
    if filter_Status:
        test1 = test1.filter(Status=filter_Status)


    unique_status = Ordertracking.objects.order_by('Status').values_list('Status', flat=True).distinct()
    unique_incharge = Ordertracking.objects.order_by('InCharger').values_list('InCharger', flat=True).distinct()



    context = {
        'form': form,
        'test1': test1,
        'filter_Status': filter_Status,
        'filter_Incharge': filter_Incharge,
        'unique_status': unique_status,
        'unique_incharge': unique_incharge,
    }
    return render(request, 'ordertracking/ordertracking.html', context)

def dropdown_status(request):
    selected_Incharge = request.GET.get('InCharger')  # Fix the case here
    print(selected_Incharge)
    test1 = Ordertracking.objects.all()

    if selected_Incharge:
        invoices = test1.filter(InCharger=selected_Incharge)  # Fix the case here
        Status_options = list(invoices.values_list('Status', flat=True).distinct())
    else:
        Status_options = []

    response_data = {
        'Status_options': Status_options,
    }

    return JsonResponse(response_data)

def clean_data(value):
    # Replace problematic characters with valid characters
    # cleaned_value = value.replace('/', '_').replace(':', '_').replace('*', '_').replace('?', '_').replace('"', '_').replace('<', '_').replace('>', '_').replace('|', '_')
    # def clean_data(value):
    # Define a regular expression pattern to match non-printable characters

    if not isinstance(value, str):
        # Handle non-string values here, e.g., by returning an empty string
        return ""
    pattern = r'[\x00-\x1F\x7F-\x9F\xAD]'

    # Use the regular expression to replace non-printable characters with an underscore
    cleaned_value = re.sub(pattern, "_", value)

    return cleaned_value    
 

def ordertracking_csv(request):
    filter_Incharge = request.GET.get('filter_Incharge')
    filter_Status = request.GET.get('filter_Status')
   

    filters = {}
    if filter_Incharge:
        filters['InCharger'] = filter_Incharge
    if filter_Status:
        filters['Status'] = filter_Status
    

    test1 = Ordertracking.objects.filter(**filters)  # Apply the filters to the queryset

    # Calculate total amounts
    # total_invoice_amount = sum(float(data.Invoiceamount) for data in test1)
    # total_outstanding_amount = sum(float(data.outstandingBalance) for data in test1 if data.outstandingBalance.strip())

    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Filtered Data"

    fields = ['ID','Branch','WES_NO','PO_NO','PO_Date', 'Client_Name','Vessel_Name','Supplier_Name','Forwarder_name','AWB_NO','Status','Status_Date','InCharger','Dolibar','Remarks']
 
    # Write the headers
    for col_num, field in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_num, value=field)
        cell.font = openpyxl.styles.Font(bold=True)


    # for row_num, data in enumerate(test1, 2):
    #     for col_num, field in enumerate(fields, 1):
    #         cleaned_value = clean_data(getattr(data, field))
    #         cell = ws.cell(row=row_num, column=col_num, value=cleaned_value)
    for row_num, data in enumerate(test1, 2):
        for col_num, field in enumerate(fields, 1):
            if field == 'ID':
                cell = ws.cell(row=row_num, column=col_num, value=getattr(data, field))  # Use the value directly
            else:
                cleaned_value = clean_data(getattr(data, field))
                cell = ws.cell(row=row_num, column=col_num, value=cleaned_value)


    file_path = "filtered_data.xlsx"
    wb.save(file_path)

    response = FileResponse(open(file_path, 'rb'), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="filtered_data.xlsx"'

    return response    

def tracking_csv(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Full Data"
    testdata_objects = Ordertracking.objects.all()

    # Define the fields for the CSV file
    fields = ['ID','Branch','WES_NO','PO_NO','PO_Date', 'Client_Name','Vessel_Name','Supplier_Name','Forwarder_name','AWB_NO','Status','Status_Date','InCharger','Dolibar','Remarks']
    
    for col_num, field in enumerate(fields, 1):
        # col_letter = get_column_letter(col_num)
        cell = ws.cell(row=1, column=col_num, value=field)
        cell.font = openpyxl.styles.Font(bold=True) 

    # Write the data rows
    for row_num, data in enumerate(testdata_objects, 2):
        for col_num, field in enumerate(fields, 1):
            cell = ws.cell(row=row_num, column=col_num, value=getattr(data, field))

    file_path = "full_data.xlsx"
    wb.save(file_path)

    response = FileResponse(open(file_path, 'rb'), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="full_data.xlsx'

    return response         

def addtracking(request):
    if request.method == "POST":
        form = OrdertrackingForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                return redirect('Logistics:ordertracking')
            except Exception as e:
                return HttpResponse(f"An error occurred: {str(e)}")
        else:
            return HttpResponse("Form is not valid.")
    else:
        form = OrdertrackingForm()

    Branch = Ordertracking.objects.exclude(Status__exact='').order_by('Branch').values_list('Branch', flat=True).distinct()
    Status = Ordertracking.objects.exclude(Status__exact='').order_by('Status').values_list('Status', flat=True).distinct()
    InCharger = Ordertracking.objects.exclude(InCharger__exact='').order_by('InCharger').values_list('InCharger', flat=True).distinct()
    Client_Name = Ordertracking.objects.exclude(Status__exact='').order_by('Client_Name').values_list('Client_Name', flat=True).distinct()
    Vessel_Name = Ordertracking.objects.exclude(Status__exact='').order_by('Vessel_Name').values_list('Vessel_Name', flat=True).distinct()
    Supplier_Name = Ordertracking.objects.exclude(InCharger__exact='').order_by('Supplier_Name').values_list('Supplier_Name', flat=True).distinct()
    Forwarder_name = Ordertracking.objects.exclude(Status__exact='').order_by('Forwarder_name').values_list('Forwarder_name', flat=True).distinct()
    Dolibar = Ordertracking.objects.exclude(InCharger__exact='').order_by('Dolibar').values_list('Dolibar', flat=True).distinct()

    last_id = Ordertracking.objects.last().ID if Ordertracking.objects.exists() else 0
    context = {'last_id': last_id + 1} 
      

    context = {
        'form': form,
        'Status_options': Status,
        'InCharger_options' : InCharger,
        'Branch_options': Branch,
        'Client_Name_options' : Client_Name,
        'Vessel_Name_options': Vessel_Name,
        'Supplier_Name_options' : Supplier_Name,
        'Forwarder_name_options': Forwarder_name,
        'Dolibar_options' : Dolibar,
        'last_id': last_id + 1

    }
    return render(request, 'ordertracking/addtracking.html', context)      

def addnewtracking(request):
    if request.method == "POST":
        form = OrdertrackingForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                return redirect('Logistics:ordertracking')
            except Exception as e:
                return HttpResponse(f"An error occurred: {str(e)}")
        else:
            return HttpResponse("Form is not valid.")
    else:
        form = OrdertrackingForm()
 
    last_id = Ordertracking.objects.last().ID if Ordertracking.objects.exists() else 0
    context = {'last_id': last_id + 1} 
      

    context = {
        'form': form,
        'last_id': last_id + 1

    }
    return render(request, 'ordertracking/addnewtracking.html', context)

def filterbtn(request):
    filter_Status = request.GET.get('filter_Status')
    filter_Incharge = request.GET.get('filter_Incharge')

    test1 = Ordertracking.objects.all().order_by('-ID')


    if filter_Status:
        test1 = test1.filter(Status=filter_Status)
    if filter_Incharge:
        test1 = test1.filter(InCharger=filter_Incharge)
   
    filtered_data = list(test1.values())    
    total_rows = len(filtered_data)
    context = {
        'test1': test1,
        'filter_Status': filter_Status,
        'filter_Incharge': filter_Incharge,
        'total_rows':int(total_rows),
    }
    # return JsonResponse({'total_rows': total_rows})
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        return JsonResponse(filtered_data, safe=False)
    
    return render(request,'ordertracking/filterbtn.html', context)

def updatetracking(request, data_id):
    data = get_object_or_404(Ordertracking, pk=data_id)
    if request.method == 'POST':
        form = OrdertrackingForm(request.POST, instance=data)

        if form.is_valid():
            form.save()
            return redirect('Logistics:ordertracking')
        else:
            print('invalid')    
    else:
        form = OrdertrackingForm(instance=data)
        
    Branch = Ordertracking.objects.exclude(Branch__exact='').order_by('Branch').values_list('Branch', flat=True).distinct()
    Status = Ordertracking.objects.exclude(Status__exact='').order_by('Status').values_list('Status', flat=True).distinct()
    InCharger = Ordertracking.objects.exclude(InCharger__exact='').order_by('InCharger').values_list('InCharger', flat=True).distinct()
    Client_Name = Ordertracking.objects.exclude(Client_Name__exact='').order_by('Client_Name').values_list('Client_Name', flat=True).distinct()
    Vessel_Name = Ordertracking.objects.exclude(Vessel_Name__exact='').order_by('Vessel_Name').values_list('Vessel_Name', flat=True).distinct()
    Supplier_Name = Ordertracking.objects.exclude(Supplier_Name__exact='').order_by('Supplier_Name').values_list('Supplier_Name', flat=True).distinct()
    Forwarder_name = Ordertracking.objects.exclude(Forwarder_name__exact='').order_by('Forwarder_name').values_list('Forwarder_name', flat=True).distinct()
    Dolibar = Ordertracking.objects.exclude(Dolibar__exact='').order_by('Dolibar').values_list('Dolibar', flat=True).distinct()

    last_id = Ordertracking.objects.last().ID if Ordertracking.objects.exists() else 0
   
      

    context = {
        'form': form,
        'data_id': data_id,
        'Status_options': Status,
        'InCharger_options' : InCharger,
        'Branch_options': Branch,
        'Client_Name_options' : Client_Name,
        'Vessel_Name_options': Vessel_Name,
        'Supplier_Name_options' : Supplier_Name,
        'Forwarder_name_options': Forwarder_name,
        'Dolibar_options' : Dolibar,
        'last_id': last_id + 1

    }
    return render(request, 'ordertracking/updatetracking.html', context)  

def order(request):

    test1 = Ordertracking.objects.all()

    print(test1)
    test1 = test1.filter(Status='Order Confirmation').order_by('-ID')

    filtered_data = list(test1.values())    
    total_rows = len(filtered_data)
    context = {
        'test1': test1,
        'total_rows':int(total_rows) ,
    }
    # return JsonResponse({'total_rows': total_rows})
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        return JsonResponse(filtered_data, safe=False)
    
    return render(request,'ordertracking/order.html', context)      

def information(request):

    test1 = Ordertracking.objects.all()
    print(test1)
    test1 = test1.filter(Status='Waiting For information').order_by('-ID')

    filtered_data = list(test1.values())    
    total_rows = len(filtered_data)
    context = {
        'test1': test1,
        'total_rows':int(total_rows) ,
    }
    # return JsonResponse({'total_rows': total_rows})
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        return JsonResponse(filtered_data, safe=False)
    
    return render(request,'ordertracking/information.html', context)     

def consignee(request):

    test1 = Ordertracking.objects.all()
    print(test1)
    test1 = test1.filter(Status='Waiting for consignee').order_by('-ID')

    filtered_data = list(test1.values())    
    total_rows = len(filtered_data)
    context = {
        'test1': test1,
        'total_rows':int(total_rows) ,
    }
    # return JsonResponse({'total_rows': total_rows})
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        return JsonResponse(filtered_data, safe=False)
    
    return render(request,'ordertracking/consignee.html', context)      

def freightcost(request):

    test1 = Ordertracking.objects.all()
    print(test1)
    test1 = test1.filter(Status='Freight cost quoted').order_by('-ID')

    filtered_data = list(test1.values())    
    total_rows = len(filtered_data)
    context = {
        'test1': test1,
        'total_rows':int(total_rows) ,
    }
    # return JsonResponse({'total_rows': total_rows})
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        return JsonResponse(filtered_data, safe=False)
    
    return render(request,'ordertracking/freightcost.html', context)      

def arrangeddispatch(request):

    test1 = Ordertracking.objects.all()
    print(test1)
    test1 = test1.filter(Status='Arranged dispatch').order_by('-ID')

    filtered_data = list(test1.values())    
    total_rows = len(filtered_data)
    context = {
        'test1': test1,
        'total_rows':int(total_rows) ,
    }
    # return JsonResponse({'total_rows': total_rows})
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        return JsonResponse(filtered_data, safe=False)
    
    return render(request,'ordertracking/arrangeddispatch.html', context)    

def dispatched(request):

    test1 = Ordertracking.objects.all()
    print(test1)
    test1 = test1.filter(Status='Dispatched').order_by('-ID')

    filtered_data = list(test1.values())    
    total_rows = len(filtered_data)
    context = {
        'test1': test1,
        'total_rows':int(total_rows) ,
    }
    # return JsonResponse({'total_rows': total_rows})
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        return JsonResponse(filtered_data, safe=False)
    
    return render(request,'ordertracking/dispatched.html', context)         

def delivered(request):

    test1 = Ordertracking.objects.all()
    print(test1)
    test1 = test1.filter(Status='Delivered').order_by('-ID')

    filtered_data = list(test1.values())    
    total_rows = len(filtered_data)
    context = {
        'test1': test1,
        'total_rows':int(total_rows) ,
    }
    # return JsonResponse({'total_rows': total_rows})
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        return JsonResponse(filtered_data, safe=False)
    
    return render(request,'ordertracking/delivered.html', context)     



def pending(request):

    test1 = Ordertracking.objects.all()
    print(test1)
    test1 = test1.filter(Status='Pending and Long Lead Time').order_by('-ID')

    filtered_data = list(test1.values())    
    total_rows = len(filtered_data)
    context = {
        'test1': test1,
        'total_rows':int(total_rows) ,
    }
    # return JsonResponse({'total_rows': total_rows})
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        return JsonResponse(filtered_data, safe=False)
    
    return render(request,'ordertracking/pending.html', context)         


def invoiced(request):

    test1 = Ordertracking.objects.all()
    print(test1)
    test1 = test1.filter(Status='Invoiced').order_by('-ID')

    filtered_data = list(test1.values())    
    total_rows = len(filtered_data)
    context = {
        'test1': test1,
        'total_rows':int(total_rows) ,
    }
    # return JsonResponse({'total_rows': total_rows})
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        return JsonResponse(filtered_data, safe=False)
    
    return render(request,'ordertracking/invoiced.html', context)         

def holdpo(request):

    test1 = Ordertracking.objects.all()
    print(test1)
    test1 = test1.filter(Status='Hold PO').order_by('-ID')

    filtered_data = list(test1.values())    
    total_rows = len(filtered_data)
    context = {
        'test1': test1,
        'total_rows':int(total_rows) ,
    }
    # return JsonResponse({'total_rows': total_rows})
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        return JsonResponse(filtered_data, safe=False)
    
    return render(request,'ordertracking/holdpo.html', context)   

def ongoing(request):

    test1 = Ordertracking.objects.all()
    print(test1)
    test1 = test1.filter(Status='Clarification ongoing').order_by('-ID')

    filtered_data = list(test1.values())    
    total_rows = len(filtered_data)
    context = {
        'test1': test1,
        'total_rows':int(total_rows) ,
    }
    # return JsonResponse({'total_rows': total_rows})
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        return JsonResponse(filtered_data, safe=False)
    
    return render(request,'ordertracking/ongoing.html', context)         

